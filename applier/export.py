import csv
import json
import html as _html
from pathlib import Path
from datetime import datetime
from rich.console import Console
from rich.table import Table
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .search.base import JobResult

console = Console()


def to_table(results: list[JobResult]) -> None:
    table = Table(title=f"Jobs found: {len(results)}", show_lines=False)
    table.add_column("#", style="dim", width=4)
    table.add_column("Title", style="bold")
    table.add_column("Company")
    table.add_column("Location")
    table.add_column("Platform", style="cyan")
    table.add_column("Contract", style="green")

    for i, r in enumerate(results, 1):
        table.add_row(str(i), r.title, r.company, r.location, r.platform, r.contract)

    console.print(table)


def to_markdown(results: list[JobResult], path: Path) -> None:
    lines = [
        f"# Job Results — {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"\n{len(results)} listings found.\n",
    ]
    for i, r in enumerate(results, 1):
        lines.append(f"{i}. **[{r.title}]({r.url})** — {r.company} | {r.location} | {r.platform} | {r.contract}")
    path.write_text("\n".join(lines), encoding="utf-8")
    console.print(f"[green]Saved {len(results)} results → {path}[/green]")


def to_csv(results: list[JobResult], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["title", "company", "location", "url", "platform", "contract"])
        w.writeheader()
        for r in results:
            w.writerow({"title": r.title, "company": r.company, "location": r.location,
                        "url": r.url, "platform": r.platform, "contract": r.contract})
    console.print(f"[green]Saved {len(results)} results → {path}[/green]")


def to_json(results: list[JobResult], path: Path) -> None:
    data = [{"title": r.title, "company": r.company, "location": r.location,
             "url": r.url, "platform": r.platform, "contract": r.contract}
            for r in results]
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    console.print(f"[green]Saved {len(results)} results → {path}[/green]")


EXCEL_HEADERS = ["Title", "Company", "Location", "Contract", "Platform", "URL",
                 "Status", "Applied Date", "Notes"]
TRACKING_COLS = {"Status", "Applied Date", "Notes"}
STATUS_OPTIONS = "To Apply / Applied / Interview / Rejected / Offer / Ignore"

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_STATUS_FILLS = {
    "Applied":   PatternFill("solid", fgColor="D6E4F0"),
    "Interview": PatternFill("solid", fgColor="D5F5E3"),
    "Offer":     PatternFill("solid", fgColor="A9DFBF"),
    "Rejected":  PatternFill("solid", fgColor="FADBD8"),
    "Ignore":    PatternFill("solid", fgColor="F2F3F4"),
}
_COL_WIDTHS = [50, 30, 25, 20, 14, 60, 14, 14, 40]


def to_excel(results: list[JobResult], path: Path) -> None:
    """Write/merge results into a tracking Excel file.

    If the file already exists, only NEW urls are appended — existing rows
    (with Status / Applied Date / Notes filled in) are never touched.
    """
    existing_urls: dict[str, int] = {}  # url → row number (1-based)
    existing_tracking: dict[str, tuple] = {}  # url → (status, date, notes)

    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        url_col = EXCEL_HEADERS.index("URL") + 1
        status_col = EXCEL_HEADERS.index("Status") + 1
        date_col = EXCEL_HEADERS.index("Applied Date") + 1
        notes_col = EXCEL_HEADERS.index("Notes") + 1
        for row_idx in range(2, ws.max_row + 1):
            url_val = ws.cell(row=row_idx, column=url_col).value
            if url_val:
                existing_urls[url_val] = row_idx
                existing_tracking[url_val] = (
                    ws.cell(row=row_idx, column=status_col).value or "",
                    ws.cell(row=row_idx, column=date_col).value or "",
                    ws.cell(row=row_idx, column=notes_col).value or "",
                )
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs"
        ws.freeze_panes = "A2"

        # Header row
        for col_idx, header in enumerate(EXCEL_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        for col_idx, width in enumerate(_COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 20

    # Append new rows
    new_count = 0
    for r in results:
        if r.url in existing_urls:
            continue
        row = [r.title, r.company, r.location, r.contract, r.platform, r.url,
               "", "", ""]
        ws.append(row)
        row_idx = ws.max_row

        # Hyperlink the URL cell
        url_col_idx = EXCEL_HEADERS.index("URL") + 1
        url_cell = ws.cell(row=row_idx, column=url_col_idx)
        url_cell.hyperlink = r.url
        url_cell.font = Font(color="0563C1", underline="single")

        new_count += 1

    # Add data validation hint in Status column header comment
    ws.cell(row=1, column=EXCEL_HEADERS.index("Status") + 1).comment = None

    wb.save(path)
    verb = "Updated" if existing_urls else "Saved"
    console.print(
        f"[green]{verb} → {path} "
        f"({new_count} new rows, {len(existing_urls)} existing preserved)[/green]"
    )
    if existing_urls:
        console.print(f"  [dim]Status options: {STATUS_OPTIONS}[/dim]")


def to_html(results: list[JobResult], path: Path, xlsx_path: Path | None = None) -> None:
    """Generate a fully self-contained HTML tracker — no server needed.

    Tracking data (status / applied_date / notes) is loaded from the Excel file
    if it exists, embedded in the HTML, and persisted in localStorage on every change.
    Re-running this overwrites the HTML but merges localStorage state back in on open.
    """
    # Pull existing tracking from Excel so we don't lose it on re-export
    tracking: dict[str, dict] = {}
    src = xlsx_path or path.with_suffix(".xlsx")
    if src.exists():
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers) if h}
        for row in ws.iter_rows(min_row=2, values_only=True):
            url = row[col.get("URL", -1)] if "URL" in col else None
            if not url:
                continue
            raw_date = row[col.get("Applied Date", -1)] if "Applied Date" in col else None
            date_str = ""
            if raw_date:
                date_str = raw_date.strftime("%Y-%m-%d") if hasattr(raw_date, "strftime") else str(raw_date)
            tracking[url] = {
                "status":       row[col.get("Status", -1)] or "",
                "applied_date": date_str,
                "notes":        row[col.get("Notes", -1)] or "",
            }
        wb.close()

    jobs_data = []
    for r in results:
        t = tracking.get(r.url, {})
        jobs_data.append({
            "title":        r.title,
            "company":      r.company,
            "location":     r.location,
            "contract":     r.contract,
            "platform":     r.platform,
            "url":          r.url,
            "tags":         r.tags if r.tags else [],
            "pulled_at":    r.pulled_at or "",
            "status":       t.get("status", ""),
            "applied_date": t.get("applied_date", ""),
            "notes":        t.get("notes", ""),
        })

    jobs_json = json.dumps(jobs_data, ensure_ascii=False)
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Job Tracker — Hamza Kritet</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:opsz,wght@14..32,300..800&display=swap" rel="stylesheet">
<style>
:root{{
  --bg:#0d1117;--s1:#161b22;--s2:#1c2330;--s3:#242f3d;
  --b1:rgba(255,255,255,.08);--b2:rgba(255,255,255,.13);
  --t1:#e6edf3;--t2:#8b949e;--t3:#484f58;
  --acc:#4f8ef7;--acc-d:rgba(79,142,247,.15);--acc-g:rgba(79,142,247,.28);
  --green:#3fb950;--green-d:rgba(63,185,80,.14);
  --teal:#39d3c3;--teal-d:rgba(57,211,195,.13);
  --amber:#d29922;--amber-d:rgba(210,153,34,.13);
  --red:#f85149;--red-d:rgba(248,81,73,.13);
  --gray:#6e7681;--gray-d:rgba(110,118,129,.12);
  --r:10px;--r-sm:6px;--r-lg:14px;
  --ease:cubic-bezier(.4,0,.2,1);
}}
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
html{{scroll-behavior:smooth}}
body{{font-family:'Inter',system-ui,sans-serif;background:var(--bg);color:var(--t1);
  min-height:100vh;font-size:13px;line-height:1.5;-webkit-font-smoothing:antialiased}}

/* Nav */
nav{{position:sticky;top:0;z-index:200;height:54px;
  display:flex;align-items:center;gap:14px;padding:0 28px;
  background:rgba(13,17,23,.82);backdrop-filter:blur(20px) saturate(1.4);
  border-bottom:1px solid var(--b1);}}
.logo{{width:28px;height:28px;border-radius:var(--r-sm);
  background:linear-gradient(135deg,var(--acc),#7c3aed);
  display:grid;place-items:center;font-size:14px;
  box-shadow:0 0 16px var(--acc-g);flex-shrink:0}}
.nav-title{{font-size:13.5px;font-weight:700;letter-spacing:-.2px}}
.nav-sub{{font-size:11px;color:var(--t3);margin-left:4px}}
.nav-r{{margin-left:auto;display:flex;align-items:center;gap:10px}}
.pill{{font-size:11px;font-weight:600;color:var(--acc);
  background:var(--acc-d);border:1px solid rgba(79,142,247,.22);
  padding:3px 12px;border-radius:100px}}
.btn-export{{display:flex;align-items:center;gap:6px;padding:6px 14px;
  background:var(--s2);border:1px solid var(--b2);border-radius:var(--r-sm);
  font:inherit;font-size:12px;color:var(--t2);cursor:pointer;
  transition:120ms var(--ease);}}
.btn-export:hover{{background:var(--s3);color:var(--t1);border-color:var(--b2)}}

/* Stats */
.stats{{background:var(--s1);border-bottom:1px solid var(--b1);
  padding:12px 28px;display:flex;gap:8px;flex-wrap:wrap}}
.sc{{display:flex;align-items:center;gap:10px;padding:9px 14px;
  background:var(--s2);border:1px solid var(--b1);border-radius:var(--r);
  cursor:pointer;transition:120ms var(--ease);user-select:none}}
.sc:hover{{background:var(--s3);border-color:var(--b2);transform:translateY(-1px)}}
.sc.on{{border-color:var(--acc);background:var(--acc-d)}}
.sc-icon{{width:30px;height:30px;border-radius:var(--r-sm);
  display:grid;place-items:center;font-size:14px;flex-shrink:0}}
.sc-n{{font-size:18px;font-weight:700;line-height:1.1;font-variant-numeric:tabular-nums}}
.sc-l{{font-size:10px;font-weight:600;color:var(--t3);text-transform:uppercase;letter-spacing:.5px;margin-top:1px}}
.c-all   .sc-icon{{background:var(--gray-d)}} .c-all   .sc-n{{color:var(--t1)}}
.c-todo  .sc-icon{{background:var(--acc-d)}}  .c-todo  .sc-n{{color:var(--acc)}}
.c-app   .sc-icon{{background:var(--teal-d)}} .c-app   .sc-n{{color:var(--teal)}}
.c-int   .sc-icon{{background:var(--green-d)}}.c-int   .sc-n{{color:var(--green)}}
.c-off   .sc-icon{{background:var(--amber-d)}}.c-off   .sc-n{{color:var(--amber)}}
.c-rej   .sc-icon{{background:var(--red-d)}}  .c-rej   .sc-n{{color:var(--red)}}

/* Toolbar */
.toolbar{{background:var(--s1);border-bottom:1px solid var(--b1);
  padding:10px 28px;display:flex;gap:10px;flex-wrap:wrap;align-items:center}}
.sw{{position:relative;flex:1;min-width:220px}}
.sw svg{{position:absolute;left:11px;top:50%;transform:translateY(-50%);
  color:var(--t3);pointer-events:none;transition:color 120ms var(--ease)}}
.sw:focus-within svg{{color:var(--acc)}}
.si{{width:100%;padding:8px 12px 8px 34px;background:var(--s2);
  border:1px solid var(--b1);border-radius:var(--r);font:inherit;font-size:13px;
  color:var(--t1);outline:none;transition:120ms var(--ease)}}
.si::placeholder{{color:var(--t3)}}
.si:focus{{border-color:var(--acc);box-shadow:0 0 0 3px var(--acc-g)}}
.fb{{padding:8px 30px 8px 12px;background:var(--s2);border:1px solid var(--b1);
  border-radius:var(--r);font:inherit;font-size:12px;color:var(--t2);
  cursor:pointer;outline:none;appearance:none;
  background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='10' viewBox='0 0 24 24' fill='none' stroke='%238b949e' stroke-width='2.5'%3E%3Cpath d='M6 9l6 6 6-6'/%3E%3C/svg%3E");
  background-repeat:no-repeat;background-position:right 9px center;
  transition:120ms var(--ease)}}
.fb:focus{{border-color:var(--acc);box-shadow:0 0 0 3px var(--acc-g);color:var(--t1)}}
.fb option{{background:var(--s2)}}
.ct{{font-size:11.5px;color:var(--t3);margin-left:auto;white-space:nowrap}}

/* Table */
.tw{{padding:16px 28px 48px}}
.tc{{background:var(--s1);border:1px solid var(--b1);border-radius:var(--r-lg);
  overflow:hidden;box-shadow:0 8px 32px rgba(0,0,0,.4)}}
table{{width:100%;border-collapse:collapse}}
thead tr{{background:var(--s2);border-bottom:1px solid var(--b1)}}
th{{padding:9px 13px;text-align:left;font-size:10px;font-weight:600;
  color:var(--t3);text-transform:uppercase;letter-spacing:.7px;
  white-space:nowrap;cursor:pointer;user-select:none;transition:color 120ms var(--ease)}}
th:hover{{color:var(--t2)}}
th.asc::after{{content:" ↑";color:var(--acc)}}
th.desc::after{{content:" ↓";color:var(--acc)}}
tbody tr{{border-bottom:1px solid var(--b1);transition:background 100ms var(--ease)}}
tbody tr:last-child{{border-bottom:none}}
tbody tr:hover{{background:rgba(255,255,255,.02)}}
tbody tr.r-Applied  {{border-left:2px solid var(--teal)}}
tbody tr.r-Interview{{border-left:2px solid var(--green)}}
tbody tr.r-Offer    {{border-left:2px solid var(--amber)}}
tbody tr.r-Rejected {{opacity:.38}}
tbody tr.r-Ignore   {{opacity:.22}}
td{{padding:10px 13px;vertical-align:middle}}
.tn{{color:var(--t3);font-size:11px;width:36px;font-variant-numeric:tabular-nums}}
.tt{{max-width:280px}}
.tt a{{color:var(--t1);text-decoration:none;font-weight:500;
  display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden;
  transition:color 120ms var(--ease)}}
.tt a:hover{{color:var(--acc)}}
.tc2{{color:var(--t2);font-size:12.5px;max-width:160px}}
.tl{{color:var(--t3);font-size:12px;white-space:nowrap}}
.tk{{color:var(--t3);font-size:11.5px}}

/* Platform */
.plat{{display:inline-flex;align-items:center;gap:5px;padding:3px 9px;
  border-radius:100px;font-size:10.5px;font-weight:600;
  border:1px solid transparent;white-space:nowrap}}
.plat::before{{content:'';width:5px;height:5px;border-radius:50%;flex-shrink:0}}
.LinkedIn     {{background:rgba(10,102,194,.18);color:#58a6ff;border-color:rgba(88,166,255,.2)}}
.LinkedIn::before{{background:#58a6ff}}
.FranceTravail{{background:rgba(248,130,74,.14);color:#ffa657;border-color:rgba(255,166,87,.2)}}
.FranceTravail::before{{background:#ffa657}}
.RemoteOK     {{background:rgba(63,185,80,.13);color:#3fb950;border-color:rgba(63,185,80,.2)}}
.RemoteOK::before{{background:#3fb950}}

/* Status select */
select.ss{{padding:4px 24px 4px 9px;border-radius:100px;font:inherit;font-size:11.5px;
  font-weight:600;cursor:pointer;outline:none;border:1px solid transparent;
  appearance:none;min-width:106px;
  background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='9' height='9' viewBox='0 0 24 24' fill='none' stroke='%238b949e' stroke-width='2.5'%3E%3Cpath d='M6 9l6 6 6-6'/%3E%3C/svg%3E");
  background-repeat:no-repeat;background-position:right 7px center;
  transition:120ms var(--ease)}}
select.ss:focus{{box-shadow:0 0 0 3px var(--acc-g)}}
select.ss option{{background:var(--s2)}}
.ss-x       {{background:var(--s2);       color:var(--t2); border-color:var(--b1)}}
.ss-Applied {{background:var(--teal-d);   color:var(--teal);  border-color:rgba(57,211,195,.25)}}
.ss-Interview{{background:var(--green-d); color:var(--green); border-color:rgba(63,185,80,.25)}}
.ss-Offer   {{background:var(--amber-d);  color:var(--amber); border-color:rgba(210,153,34,.25)}}
.ss-Rejected{{background:var(--red-d);    color:var(--red);   border-color:rgba(248,81,73,.25)}}
.ss-Ignore  {{background:var(--gray-d);   color:var(--gray);  border-color:var(--b1)}}

/* Inputs */
.ci{{width:100%;padding:4px 7px;background:transparent;border:1px solid transparent;
  border-radius:var(--r-sm);font:inherit;font-size:12px;color:var(--t2);
  outline:none;transition:120ms var(--ease)}}
.ci::placeholder{{color:var(--t3)}}
.ci:hover{{border-color:var(--b2);background:var(--s2)}}
.ci:focus{{border-color:var(--acc);background:var(--s2);
  box-shadow:0 0 0 3px var(--acc-g);color:var(--t1)}}

/* Empty */
.empty{{display:none;flex-direction:column;align-items:center;
  gap:12px;padding:64px 24px;color:var(--t3)}}
.empty svg{{opacity:.25}}
.empty strong{{font-size:15px;color:var(--t2)}}
.empty span{{font-size:12.5px}}

/* Toast */
#toast{{position:fixed;bottom:22px;right:22px;display:flex;align-items:center;gap:8px;
  background:var(--s2);border:1px solid var(--b2);color:var(--t1);
  padding:9px 15px;border-radius:var(--r);font-size:13px;font-weight:500;
  box-shadow:0 8px 32px rgba(0,0,0,.5);transform:translateY(10px) scale(.97);
  opacity:0;transition:opacity 180ms var(--ease),transform 180ms var(--ease);
  pointer-events:none;z-index:999}}
#toast.on{{opacity:1;transform:translateY(0) scale(1)}}
.tdot{{width:7px;height:7px;border-radius:50%;background:var(--green);
  box-shadow:0 0 6px var(--green)}}

/* Tag chips */
.chip{{display:inline-block;padding:2px 7px;border-radius:100px;
  font-size:10px;font-weight:600;margin:1px 2px 1px 0;white-space:nowrap;
  background:var(--s3);color:var(--t2);border:1px solid var(--b1)}}
.chip-IT{{background:rgba(79,142,247,.15);color:#79b8ff;border-color:rgba(79,142,247,.25)}}
.chip-Security{{background:rgba(248,81,73,.13);color:#ffa198;border-color:rgba(248,81,73,.22)}}
.chip-Sales{{background:rgba(57,211,195,.13);color:var(--teal);border-color:rgba(57,211,195,.22)}}
.chip-Marketing{{background:rgba(210,153,34,.13);color:var(--amber);border-color:rgba(210,153,34,.22)}}
.chip-Finance{{background:rgba(63,185,80,.13);color:var(--green);border-color:rgba(63,185,80,.22)}}
.chip-Consulting{{background:rgba(139,94,249,.15);color:#c9b1ff;border-color:rgba(139,94,249,.25)}}
.chip-Remote{{background:rgba(63,185,80,.1);color:#56d364;border-color:rgba(63,185,80,.2)}}
.chip-Other{{background:var(--gray-d);color:var(--gray);border-color:var(--b1)}}
.pulled{{font-size:10px;color:var(--t3);margin-top:3px}}

.hidden{{display:none!important}}
::-webkit-scrollbar{{width:5px;height:5px}}
::-webkit-scrollbar-thumb{{background:var(--s3);border-radius:3px}}
</style>
</head>
<body>

<nav>
  <div class="logo">🎯</div>
  <span class="nav-title">Job Tracker</span>
  <span class="nav-sub">Generated {generated}</span>
  <div class="nav-r">
    <span class="pill" id="total-pill">{len(results)} jobs</span>
    <button class="btn-export" onclick="exportCSV()">
      <svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2">
        <path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4M7 10l5 5 5-5M12 15V3"/>
      </svg>
      Export CSV
    </button>
  </div>
</nav>

<div class="stats">
  <div class="sc c-all on" onclick="statFilter('')" id="sc-all">
    <div class="sc-icon">📋</div>
    <div><div class="sc-n" id="s-all">0</div><div class="sc-l">Total</div></div>
  </div>
  <div class="sc c-todo" onclick="statFilter('To Apply')" id="sc-todo">
    <div class="sc-icon">📌</div>
    <div><div class="sc-n" id="s-todo">0</div><div class="sc-l">To Apply</div></div>
  </div>
  <div class="sc c-app" onclick="statFilter('Applied')" id="sc-app">
    <div class="sc-icon">✉️</div>
    <div><div class="sc-n" id="s-app">0</div><div class="sc-l">Applied</div></div>
  </div>
  <div class="sc c-int" onclick="statFilter('Interview')" id="sc-int">
    <div class="sc-icon">🗓</div>
    <div><div class="sc-n" id="s-int">0</div><div class="sc-l">Interview</div></div>
  </div>
  <div class="sc c-off" onclick="statFilter('Offer')" id="sc-off">
    <div class="sc-icon">🏆</div>
    <div><div class="sc-n" id="s-off">0</div><div class="sc-l">Offer</div></div>
  </div>
  <div class="sc c-rej" onclick="statFilter('Rejected')" id="sc-rej">
    <div class="sc-icon">✕</div>
    <div><div class="sc-n" id="s-rej">0</div><div class="sc-l">Rejected</div></div>
  </div>
</div>

<div class="toolbar">
  <div class="sw">
    <svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.3">
      <circle cx="11" cy="11" r="8"/><path d="M21 21l-4.35-4.35"/>
    </svg>
    <input class="si" type="text" id="q" placeholder="Search title, company, location…" oninput="filter()">
  </div>
  <select class="fb" id="fp" onchange="filter()">
    <option value="">All platforms</option>
    <option>LinkedIn</option>
    <option>France Travail</option>
    <option>RemoteOK</option>
  </select>
  <select class="fb" id="ft" onchange="filter()">
    <option value="">All tags</option>
    <option>IT</option>
    <option>Security</option>
    <option>Sales</option>
    <option>Marketing</option>
    <option>Finance</option>
    <option>Consulting</option>
    <option>HR</option>
    <option>Product</option>
    <option>Operations</option>
    <option>Customer</option>
    <option>Remote</option>
    <option>Other</option>
  </select>
  <span class="ct" id="ct"></span>
</div>

<div class="tw">
  <div class="tc">
    <div style="overflow-x:auto">
      <table>
        <thead><tr>
          <th onclick="sort(0)">#</th>
          <th onclick="sort(1)">Position</th>
          <th onclick="sort(2)">Company</th>
          <th onclick="sort(3)">Location</th>
          <th onclick="sort(4)">Contract</th>
          <th onclick="sort(5)">Platform</th>
          <th>Tags</th>
          <th>Status</th>
          <th>Applied date</th>
          <th>Notes</th>
        </tr></thead>
        <tbody id="tbody"></tbody>
      </table>
    </div>
    <div class="empty" id="empty">
      <svg width="48" height="48" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.3">
        <circle cx="11" cy="11" r="8"/><path d="M21 21l-4.35-4.35"/>
      </svg>
      <strong>Nothing here</strong>
      <span>Try a different search or clear the filters</span>
    </div>
  </div>
</div>

<div id="toast"><span class="tdot"></span><span id="tmsg">Saved</span></div>

<script>
const JOBS = {jobs_json};
const LS_KEY = 'applier_tracking_v1';
let activeStatus = '', sortCol = -1, sortDir = 1;

// Load tracking overrides from localStorage (changes made in browser)
function loadTracking() {{
  try {{ return JSON.parse(localStorage.getItem(LS_KEY) || '{{}}'); }} catch {{ return {{}}; }}
}}
function saveTracking(url, field, value) {{
  const t = loadTracking();
  if (!t[url]) t[url] = {{}};
  t[url][field] = value;
  localStorage.setItem(LS_KEY, JSON.stringify(t));
}}

// Merge localStorage overrides into jobs array on startup
const tracking = loadTracking();
const jobs = JOBS.map(j => ({{
  ...j,
  ...(tracking[j.url] || {{}}),
}}));

function escAttr(s) {{
  return String(s || '').replace(/&/g,'&amp;').replace(/"/g,'&quot;').replace(/</g,'&lt;');
}}
function escHtml(s) {{
  return String(s || '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}}

function renderRows() {{
  const tbody = document.getElementById('tbody');
  tbody.innerHTML = jobs.map((j,i) => {{
    const st = j.status || '';
    const plat = j.platform.replace(' ','');
    const tags = Array.isArray(j.tags) ? j.tags : [];
    const tagChips = tags.map(t => `<span class="chip chip-${{t}}">${{t}}</span>`).join('');
    const pulled = j.pulled_at ? `<div class="pulled">Found ${{j.pulled_at}}</div>` : '';
    return `<tr class="jr r-${{escAttr(st)}}"
      data-i="${{i}}"
      data-title="${{escAttr(j.title.toLowerCase())}}"
      data-company="${{escAttr(j.company.toLowerCase())}}"
      data-location="${{escAttr(j.location.toLowerCase())}}"
      data-platform="${{escAttr(j.platform)}}"
      data-tags="${{escAttr(tags.join(','))}}"
      data-status="${{escAttr(st)}}"
      data-url="${{escAttr(j.url)}}">
      <td class="tn">${{i+1}}</td>
      <td class="tt">
        <a href="${{escAttr(j.url)}}" target="_blank" rel="noopener">${{escHtml(j.title)}}</a>
        ${{pulled}}
      </td>
      <td class="tc2">${{escHtml(j.company)}}</td>
      <td class="tl">${{escHtml(j.location)}}</td>
      <td class="tk">${{escHtml(j.contract)}}</td>
      <td><span class="plat ${{plat}}">${{escHtml(j.platform)}}</span></td>
      <td style="white-space:nowrap">${{tagChips || '<span class="chip chip-Other">Other</span>'}}</td>
      <td><select class="ss ss-${{escAttr(st)||'x'}}" onchange="update(this,'status')">
        <option value="">To Apply</option>
        ${{['Applied','Interview','Offer','Rejected','Ignore'].map(s =>
          `<option${{st===s?' selected':''}}>${{s}}</option>`).join('')}}
      </select></td>
      <td><input class="ci" type="date" value="${{escAttr(j.applied_date)}}"
           onchange="update(this,'applied_date')"></td>
      <td><input class="ci" type="text" value="${{escAttr(j.notes)}}"
           placeholder="Add a note…" onblur="update(this,'notes')"></td>
    </tr>`;
  }}).join('');
  filter();
}}

function update(el, field) {{
  const row = el.closest('tr');
  const url = row.dataset.url;
  const value = el.value;
  saveTracking(url, field, value);
  // Update in-memory
  const idx = parseInt(row.dataset.i);
  jobs[idx][field] = value;
  if (field === 'status') {{
    row.className = [...row.classList].filter(c => !c.startsWith('r-')).join(' ')
                  + (value ? ' r-' + value : '');
    row.dataset.status = value;
    el.className = 'ss ss-' + (value || 'x');
    filter();
  }}
  toast('Saved');
}}

function statFilter(s) {{
  activeStatus = s;
  document.querySelectorAll('.sc').forEach(c => c.classList.remove('on'));
  const map = {{'':'sc-all','To Apply':'sc-todo','Applied':'sc-app',
                'Interview':'sc-int','Offer':'sc-off','Rejected':'sc-rej'}};
  document.getElementById(map[s] || 'sc-all')?.classList.add('on');
  filter();
}}

function filter() {{
  const q    = document.getElementById('q').value.toLowerCase();
  const plat = document.getElementById('fp').value;
  const tag  = document.getElementById('ft').value;
  const rows = document.querySelectorAll('.jr');
  let v = 0;
  rows.forEach(r => {{
    const txt = r.dataset.title + ' ' + r.dataset.company + ' ' + r.dataset.location;
    const rs  = r.dataset.status || 'To Apply';
    const ok  = (!q    || txt.includes(q))
             && (!plat || r.dataset.platform === plat)
             && (!tag  || r.dataset.tags.split(',').includes(tag))
             && (!activeStatus || rs === activeStatus
                || (activeStatus==='To Apply' && !r.dataset.status));
    r.classList.toggle('hidden', !ok);
    if (ok) v++;
  }});
  document.getElementById('ct').textContent = v === rows.length
    ? `${{rows.length}} results` : `${{v}} of ${{rows.length}}`;
  document.getElementById('empty').style.display = v===0 ? 'flex' : 'none';
  updateStats();
}}

function updateStats() {{
  const rows = document.querySelectorAll('.jr');
  const c = {{todo:0,Applied:0,Interview:0,Offer:0,Rejected:0}};
  rows.forEach(r => {{
    const s = r.dataset.status;
    if (!s||s==='To Apply') c.todo++;
    else if(c[s]!==undefined) c[s]++;
  }});
  document.getElementById('s-all').textContent  = rows.length;
  document.getElementById('s-todo').textContent = c.todo;
  document.getElementById('s-app').textContent  = c.Applied;
  document.getElementById('s-int').textContent  = c.Interview;
  document.getElementById('s-off').textContent  = c.Offer;
  document.getElementById('s-rej').textContent  = c.Rejected;
}}

function sort(col) {{
  const tbody = document.getElementById('tbody');
  const rows  = [...tbody.querySelectorAll('tr.jr')];
  sortDir = sortCol===col ? sortDir*-1 : 1; sortCol=col;
  rows.sort((a,b) => {{
    const av=a.cells[col]?.textContent.trim().toLowerCase()||'';
    const bv=b.cells[col]?.textContent.trim().toLowerCase()||'';
    return av<bv?-sortDir:av>bv?sortDir:0;
  }});
  rows.forEach(r=>tbody.appendChild(r));
  document.querySelectorAll('th').forEach((th,i)=>{{
    th.classList.remove('asc','desc');
    if(i===col) th.classList.add(sortDir===1?'asc':'desc');
  }});
}}

function exportCSV() {{
  const rows = document.querySelectorAll('.jr:not(.hidden)');
  const lines = [['Title','Company','Location','Contract','Platform','URL','Status','Applied Date','Notes'].join(',')];
  rows.forEach(r => {{
    const i = parseInt(r.dataset.i);
    const j = jobs[i];
    lines.push([j.title,j.company,j.location,j.contract,j.platform,j.url,
                j.status,j.applied_date,j.notes]
      .map(v=>`"${{String(v||'').replace(/"/g,'""')}}"`)
      .join(','));
  }});
  const blob = new Blob([lines.join('\\n')], {{type:'text/csv'}});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = 'hamza_jobs_export.csv';
  a.click();
}}

let _tt;
function toast(msg) {{
  const t=document.getElementById('toast');
  document.getElementById('tmsg').textContent=msg;
  t.classList.add('on');
  clearTimeout(_tt);
  _tt=setTimeout(()=>t.classList.remove('on'),1800);
}}

renderRows();
</script>
</body>
</html>"""

    path.write_text(html_content, encoding="utf-8")
    console.print(f"[green]Saved standalone HTML → {path} ({len(results)} jobs)[/green]")
    console.print(f"  [dim]Open with: open {path}[/dim]")
